"""
Excel generation module for HealthPath Agent.
Generates appointment itineraries in Excel format with multiple sheets.
"""

import os
from datetime import datetime
from typing import List, Dict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def generate_excel_document(recommendations: List[Dict], task_params: Dict, output_path: str):
    """
    Generate an Excel document with appointment recommendations.

    Args:
        recommendations: List of ranked recommendations
        task_params: Original task parameters
        output_path: Path to save the Excel file
    """

    if not OPENPYXL_AVAILABLE:
        # Fallback: generate JSON-based Excel representation
        generate_json_excel(recommendations, task_params, output_path)
        return

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Recommendations
    ws_rec = wb.create_sheet("推荐方案", 0)
    add_recommendations_sheet(ws_rec, recommendations, task_params)

    # Sheet 2: Travel Details
    if recommendations:
        ws_travel = wb.create_sheet("交通详情", 1)
        add_travel_sheet(ws_travel, recommendations)

    # Sheet 3: Summary
    ws_summary = wb.create_sheet("摘要", 2)
    add_summary_sheet(ws_summary, recommendations, task_params)

    # Save workbook
    wb.save(output_path)


def add_recommendations_sheet(ws, recommendations: List[Dict], task_params: Dict):
    """Add recommendations sheet to workbook"""

    # Header style
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data style
    data_font = Font(size=11)
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = "医旅全景路书 - 推荐方案"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')

    # Metadata
    ws['A2'] = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'] = f"科室：{task_params.get('department', '未指定')}"
    ws['A4'] = f"症状：{task_params.get('symptom', '未指定')}"

    # Headers
    headers = ["排名", "医院名称", "医生", "职称", "挂号时间", "挂号费(元)", "排队(分钟)", "评分"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Data rows
    for row_idx, rec in enumerate(recommendations, 7):
        data = [
            rec['rank'],
            rec['hospital_name'],
            rec['doctor_name'],
            rec['doctor_title'],
            rec['appointment_time'],
            rec['total_cost'],
            rec['queue_estimate_min'],
            rec['score']
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10


def add_travel_sheet(ws, recommendations: List[Dict]):
    """Add travel details sheet to workbook"""

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(size=11)
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = "交通与距离详情"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    # Headers
    headers = ["医院名称", "距离(公里)", "交通时间(分钟)", "挂号时间", "建议出发时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Data rows
    for row_idx, rec in enumerate(recommendations, 4):
        # Calculate suggested departure time (30 min before appointment)
        suggested_departure = f"提前 {rec['total_travel_time_min'] + 30} 分钟出发"

        data = [
            rec['hospital_name'],
            rec['distance_km'],
            rec['total_travel_time_min'],
            rec['appointment_time'],
            suggested_departure
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18


def add_summary_sheet(ws, recommendations: List[Dict], task_params: Dict):
    """Add summary sheet to workbook"""

    title_font = Font(bold=True, size=12)
    normal_font = Font(size=11)

    ws['A1'] = "摘要"
    ws['A1'].font = Font(bold=True, size=14)

    row = 3
    ws[f'A{row}'] = "就医需求"
    ws[f'A{row}'].font = title_font
    row += 1

    ws[f'A{row}'] = f"科室：{task_params.get('department', '未指定')}"
    ws[f'A{row}'].font = normal_font
    row += 1

    ws[f'A{row}'] = f"症状：{task_params.get('symptom', '未指定')}"
    ws[f'A{row}'].font = normal_font
    row += 1

    ws[f'A{row}'] = f"时间要求：{task_params.get('time_window', '本周')}"
    ws[f'A{row}'].font = normal_font
    row += 2

    ws[f'A{row}'] = "推荐方案"
    ws[f'A{row}'].font = title_font
    row += 1

    if recommendations:
        best_rec = recommendations[0]
        ws[f'A{row}'] = f"最优方案：{best_rec['hospital_name']}"
        ws[f'A{row}'].font = normal_font
        row += 1

        ws[f'A{row}'] = f"医生：{best_rec['doctor_name']}（{best_rec['doctor_title']}）"
        ws[f'A{row}'].font = normal_font
        row += 1

        ws[f'A{row}'] = f"时间：{best_rec['appointment_time']}"
        ws[f'A{row}'].font = normal_font
        row += 1

        ws[f'A{row}'] = f"费用：{best_rec['total_cost']} 元"
        ws[f'A{row}'].font = normal_font
        row += 1

        ws[f'A{row}'] = f"推荐理由：{best_rec['reason']}"
        ws[f'A{row}'].font = normal_font
    else:
        ws[f'A{row}'] = "未找到匹配的号源"
        ws[f'A{row}'].font = normal_font

    ws.column_dimensions['A'].width = 50


def generate_json_excel(recommendations: List[Dict], task_params: Dict, output_path: str):
    """
    Fallback: Generate JSON-based Excel representation (saved as JSON file with .xlsx extension).
    """

    import json

    content = {
        "metadata": {
            "title": "医旅全景路书",
            "generated_at": datetime.now().isoformat(),
            "department": task_params.get("department"),
            "symptom": task_params.get("symptom")
        },
        "recommendations": recommendations,
        "task_parameters": task_params,
        "sheets": {
            "recommendations": [
                {
                    "rank": rec["rank"],
                    "hospital": rec["hospital_name"],
                    "doctor": rec["doctor_name"],
                    "title": rec["doctor_title"],
                    "time": rec["appointment_time"],
                    "fee": rec["total_cost"],
                    "queue_min": rec["queue_estimate_min"],
                    "score": rec["score"],
                    "reason": rec["reason"]
                }
                for rec in recommendations
            ],
            "travel": [
                {
                    "hospital": rec["hospital_name"],
                    "distance_km": rec["distance_km"],
                    "travel_time_min": rec["total_travel_time_min"],
                    "appointment_time": rec["appointment_time"],
                    "suggested_departure": f"提前 {rec['total_travel_time_min'] + 30} 分钟出发"
                }
                for rec in recommendations
            ]
        }
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(content, f, ensure_ascii=False, indent=2)
